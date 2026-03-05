"""视频时长结果 Excel 导出。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .models import Task


class DurationExcelExporter:
    """导出 <=15s 与 >15s/探测失败 两份 Excel。"""

    SHORT_HEADERS = ("pid", "原始链接", "视频时长")
    LONG_FAILED_HEADERS = ("pid", "原始链接", "视频时长", "探测结果")

    def export_dual(
        self,
        tasks: list[Task],
        short_output_path: str,
        long_failed_output_path: str,
    ) -> None:
        self._export_short(tasks=tasks, output_path=short_output_path)
        self._export_long_failed(tasks=tasks, output_path=long_failed_output_path)

    def _export_short(self, tasks: list[Task], output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        for idx, header in enumerate(self.SHORT_HEADERS, start=1):
            ws.cell(row=1, column=idx).value = header

        row = 2
        for task in tasks:
            if task.duration_check_bucket != "le_15":
                continue
            ws.cell(row=row, column=1).value = task.pid
            ws.cell(row=row, column=2).value = task.original_link
            ws.cell(row=row, column=3).value = task.video_duration_seconds
            row += 1

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))

    def _export_long_failed(self, tasks: list[Task], output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        for idx, header in enumerate(self.LONG_FAILED_HEADERS, start=1):
            ws.cell(row=1, column=idx).value = header

        row = 2
        for task in tasks:
            if task.duration_check_bucket not in {"gt_15", "failed"}:
                continue

            result = "时长>15s"
            if task.duration_check_bucket == "failed":
                detail = (task.error_message or "").strip() or "未知错误"
                result = f"探测失败: {detail}"

            ws.cell(row=row, column=1).value = task.pid
            ws.cell(row=row, column=2).value = task.original_link
            ws.cell(row=row, column=3).value = task.video_duration_seconds
            ws.cell(row=row, column=4).value = result
            row += 1

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))

    @staticmethod
    def generate_filenames(now: datetime | None = None) -> tuple[str, str]:
        dt = now or datetime.now()
        ts = dt.strftime("%Y%m%d%H%M%S")
        return (
            f"video2prompt-duration-le15-{ts}.xlsx",
            f"video2prompt-duration-gt15-or-failed-{ts}.xlsx",
        )
