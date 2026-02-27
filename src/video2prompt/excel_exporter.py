"""Excel 导出。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import Task, TaskState
from .review_result import extract_can_translate


class ExcelExporter:
    """基于模板导出 Product ID/能否翻译/Prompt。"""

    DOUYIN_LINK_HEADER = "抖音链接"
    CAN_TRANSLATE_HEADER = "能否翻译"

    def __init__(self, template_path: str = "docs/product_prompt_template.xlsx"):
        self.template_path = template_path

    def export(self, tasks: list[Task], output_path: str) -> None:
        template = Path(self.template_path)
        if not template.exists():
            raise FileNotFoundError(f"Excel 模板不存在: {self.template_path}")

        wb = load_workbook(self.template_path)
        ws = wb.active

        headers = {
            str(ws.cell(row=1, column=col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col).value is not None
        }
        pid_col = headers.get("Product ID")
        prompt_col = headers.get("Prompt")
        if pid_col is None or prompt_col is None:
            raise ValueError("模板缺少 Product ID 或 Prompt 列")

        can_translate_col = headers.get(self.CAN_TRANSLATE_HEADER)
        if can_translate_col is None:
            can_translate_col = ws.max_column + 1
            ws.cell(row=1, column=can_translate_col).value = self.CAN_TRANSLATE_HEADER

        link_col = headers.get(self.DOUYIN_LINK_HEADER)
        if link_col is None:
            link_col = ws.max_column + 1
            ws.cell(row=1, column=link_col).value = self.DOUYIN_LINK_HEADER

        row = 2
        for task in tasks:
            if task.state not in {TaskState.COMPLETED, TaskState.FAILED, TaskState.CANCELLED} and not task.cache_hit:
                continue
            if not task.gemini_output:
                continue
            ws.cell(row=row, column=pid_col).value = task.pid
            ws.cell(row=row, column=can_translate_col).value = task.can_translate or extract_can_translate(task.gemini_output)
            ws.cell(row=row, column=prompt_col).value = task.gemini_output
            ws.cell(row=row, column=link_col).value = task.original_link
            row += 1

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    @staticmethod
    def generate_filename() -> str:
        return f"video2prompt-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
