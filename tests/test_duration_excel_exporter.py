from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from video2prompt.duration_excel_exporter import DurationExcelExporter
from video2prompt.models import Task


def test_export_dual_split_and_reason(tmp_path: Path) -> None:
    short_output = tmp_path / "short.xlsx"
    long_failed_output = tmp_path / "long_failed.xlsx"
    exporter = DurationExcelExporter()
    tasks = [
        Task(
            pid="p-001",
            original_link="https://www.douyin.com/video/1",
            video_duration_seconds=14.2,
            duration_check_bucket="le_15",
        ),
        Task(
            pid="p-002",
            original_link="https://www.douyin.com/video/2",
            video_duration_seconds=15.0,
            duration_check_bucket="le_15",
        ),
        Task(
            pid="p-003",
            original_link="https://www.douyin.com/video/3",
            video_duration_seconds=18.6,
            duration_check_bucket="gt_15",
        ),
        Task(
            pid="p-004",
            original_link="https://www.douyin.com/video/4",
            duration_check_bucket="failed",
            error_message="解析重试耗尽: timeout",
        ),
    ]

    exporter.export_dual(tasks=tasks, short_output_path=str(short_output), long_failed_output_path=str(long_failed_output))

    short_ws = load_workbook(short_output).active
    assert short_ws.cell(row=1, column=1).value == "pid"
    assert short_ws.cell(row=1, column=2).value == "原始链接"
    assert short_ws.cell(row=1, column=3).value == "视频时长"
    assert short_ws.max_row == 3
    assert short_ws.cell(row=2, column=1).value == "p-001"
    assert short_ws.cell(row=2, column=3).value == 14.2
    assert short_ws.cell(row=3, column=1).value == "p-002"
    assert short_ws.cell(row=3, column=3).value == 15.0

    long_ws = load_workbook(long_failed_output).active
    assert long_ws.cell(row=1, column=1).value == "pid"
    assert long_ws.cell(row=1, column=2).value == "原始链接"
    assert long_ws.cell(row=1, column=3).value == "视频时长"
    assert long_ws.cell(row=1, column=4).value == "探测结果"
    assert long_ws.max_row == 3
    assert long_ws.cell(row=2, column=1).value == "p-003"
    assert long_ws.cell(row=2, column=4).value == "时长>15s"
    assert long_ws.cell(row=3, column=1).value == "p-004"
    assert long_ws.cell(row=3, column=4).value == "探测失败: 解析重试耗尽: timeout"


def test_export_dual_generates_empty_files_with_headers(tmp_path: Path) -> None:
    short_output = tmp_path / "short.xlsx"
    long_failed_output = tmp_path / "long_failed.xlsx"
    exporter = DurationExcelExporter()
    tasks = [Task(pid="p-100", original_link="https://www.douyin.com/video/100", video_duration_seconds=20.0, duration_check_bucket="gt_15")]

    exporter.export_dual(tasks=tasks, short_output_path=str(short_output), long_failed_output_path=str(long_failed_output))

    short_ws = load_workbook(short_output).active
    assert short_ws.max_row == 1
    assert short_ws.cell(row=1, column=1).value == "pid"
    assert short_ws.cell(row=1, column=2).value == "原始链接"
    assert short_ws.cell(row=1, column=3).value == "视频时长"

    long_ws = load_workbook(long_failed_output).active
    assert long_ws.max_row == 2
    assert long_ws.cell(row=2, column=1).value == "p-100"
    assert long_ws.cell(row=2, column=4).value == "时长>15s"
