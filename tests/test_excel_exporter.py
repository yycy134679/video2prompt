from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from video2prompt.excel_exporter import ExcelExporter
from video2prompt.models import Task, TaskState


def test_export_auto_add_douyin_link_column(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Product ID"
    ws.cell(row=1, column=2).value = "Prompt"
    wb.save(template_path)

    exporter = ExcelExporter(template_path=str(template_path))
    tasks = [
        Task(
            pid="p-001",
            original_link="https://www.douyin.com/video/1",
            state=TaskState.COMPLETED,
            can_translate="能",
            model_output="提示词A",
        )
    ]
    exporter.export(tasks=tasks, output_path=str(output_path))

    result = load_workbook(output_path).active
    assert result.cell(row=1, column=1).value == "Product ID"
    assert result.cell(row=1, column=2).value == "Prompt"
    assert result.cell(row=1, column=3).value == "能否翻译"
    assert result.cell(row=1, column=4).value == "抖音链接"
    assert result.cell(row=2, column=1).value == "p-001"
    assert result.cell(row=2, column=2).value == "提示词A"
    assert result.cell(row=2, column=3).value == "能"
    assert result.cell(row=2, column=4).value == "https://www.douyin.com/video/1"


def test_export_use_existing_douyin_link_column(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Product ID"
    ws.cell(row=1, column=2).value = "抖音链接"
    ws.cell(row=1, column=3).value = "Prompt"
    wb.save(template_path)

    exporter = ExcelExporter(template_path=str(template_path))
    tasks = [
        Task(
            pid="p-002",
            original_link="https://www.douyin.com/video/2",
            state=TaskState.COMPLETED,
            can_translate="不能",
            model_output="提示词B",
        )
    ]
    exporter.export(tasks=tasks, output_path=str(output_path))

    result = load_workbook(output_path).active
    assert result.max_column == 4
    assert result.cell(row=1, column=2).value == "抖音链接"
    assert result.cell(row=1, column=4).value == "能否翻译"
    assert result.cell(row=2, column=1).value == "p-002"
    assert result.cell(row=2, column=2).value == "https://www.douyin.com/video/2"
    assert result.cell(row=2, column=3).value == "提示词B"
    assert result.cell(row=2, column=4).value == "不能"


def test_export_add_category_column_when_enabled(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Product ID"
    ws.cell(row=1, column=2).value = "Prompt"
    wb.save(template_path)

    exporter = ExcelExporter(template_path=str(template_path))
    tasks = [
        Task(
            pid="p-003",
            original_link="https://www.douyin.com/video/3",
            category="服饰",
            state=TaskState.COMPLETED,
            can_translate="能",
            model_output="提示词C",
        )
    ]
    exporter.export(tasks=tasks, output_path=str(output_path), include_category=True)

    result = load_workbook(output_path).active
    assert result.cell(row=1, column=5).value == "类目"
    assert result.cell(row=2, column=5).value == "服饰"
